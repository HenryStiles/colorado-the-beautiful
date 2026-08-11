# merge_spreadsheets.py
import os
import subprocess
import openpyxl

MASTER_PATH = "/Users/henrys/Library/Mobile Documents/com~apple~CloudDocs/Downloads/Outreach list (1).xlsx"
LOCAL_PATH = "/Users/henrys/source/colorado_the_beautiful/Outreach list.xlsx"

# The replacement values for Joseph Novosad (using the new master sheet schema)
# Column 1: Name of Submitter
# Column 2: Place Name
# Column 3: Place City
# Column 4: Filename of Photo
# Column 5: Story
# Column 6: Category
JOSEPH_DATA = {
    "name": "Joseph Novosad",
    "place": "Rattlesnake Canyon Arches",
    "city": "Fruita, CO",
    "filename": "Joseph_Novosad_Rattlesnake_Canyon_Arches_h.jpg",
    "story": "Rattlesnake Canyon Arches—hiking to explore McInnis Canyons National Conservation Area",
    "category": "Friends and family"
}

def main():
    if not os.path.exists(MASTER_PATH):
        print(f"Error: Master spreadsheet not found at {MASTER_PATH}")
        return

    print(f"Loading master sheet: {MASTER_PATH}...")
    wb = openpyxl.load_workbook(MASTER_PATH)
    ws = wb['SUBMISSION Yeses']

    found_row = None
    # Search for Henry's Maroon Bells row
    # Col 1 is Name of Submitter, Col 2 is Place Name
    for r_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=r_idx, column=1).value
        place = ws.cell(row=r_idx, column=2).value
        
        if name and "henry" in str(name).lower() and place and "maroon" in str(place).lower():
            found_row = r_idx
            break

    if found_row:
        print(f"Found Henry's Maroon Bells photo at row {found_row}. Replacing with Joseph Novosad's submission...")
        ws.cell(row=found_row, column=1, value=JOSEPH_DATA["name"])
        ws.cell(row=found_row, column=2, value=JOSEPH_DATA["place"])
        ws.cell(row=found_row, column=3, value=JOSEPH_DATA["city"])
        ws.cell(row=found_row, column=4, value=JOSEPH_DATA["filename"])
        ws.cell(row=found_row, column=5, value=JOSEPH_DATA["story"])
        ws.cell(row=found_row, column=6, value=JOSEPH_DATA["category"])
    else:
        print("Henry Stiles' Maroon Bells photo not found in the new master sheet. Checking if Joseph Novosad already exists...")
        # Check if Joseph is already there
        already_has_joseph = False
        for r_idx in range(2, ws.max_row + 1):
            name = ws.cell(row=r_idx, column=1).value
            place = ws.cell(row=r_idx, column=2).value
            if name and "joseph novosad" in str(name).lower():
                already_has_joseph = True
                print(f"Joseph Novosad already exists in the sheet at row {r_idx}. Overwriting with clean data...")
                ws.cell(row=r_idx, column=1, value=JOSEPH_DATA["name"])
                ws.cell(row=r_idx, column=2, value=JOSEPH_DATA["place"])
                ws.cell(row=r_idx, column=3, value=JOSEPH_DATA["city"])
                ws.cell(row=r_idx, column=4, value=JOSEPH_DATA["filename"])
                ws.cell(row=r_idx, column=5, value=JOSEPH_DATA["story"])
                ws.cell(row=r_idx, column=6, value=JOSEPH_DATA["category"])
                break
        
        if not already_has_joseph:
            # Find the actual next row of active data (max_row can return empty formatted cells)
            next_row = 2
            while ws.cell(row=next_row, column=1).value is not None:
                next_row += 1
                
            print(f"Appending Joseph Novosad's photo as a new entry at row {next_row}...")
            ws.cell(row=next_row, column=1, value=JOSEPH_DATA["name"])
            ws.cell(row=next_row, column=2, value=JOSEPH_DATA["place"])
            ws.cell(row=next_row, column=3, value=JOSEPH_DATA["city"])
            ws.cell(row=next_row, column=4, value=JOSEPH_DATA["filename"])
            ws.cell(row=next_row, column=5, value=JOSEPH_DATA["story"])
            ws.cell(row=next_row, column=6, value=JOSEPH_DATA["category"])

    # Save to the local repository path
    print(f"Saving merged sheet to local path: {LOCAL_PATH}...")
    wb.save(LOCAL_PATH)
    print("  -> Success!")

    # Re-run pipeline scripts
    print("\nRe-running auto-categorization script...")
    try:
        subprocess.run(["python3", "add_category_column.py"], check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error running category classifier: {e}")
        return

    print("Re-running Category Statistics generator...")
    try:
        subprocess.run(["python3", "add_stats_sheet.py"], check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error running stats generator: {e}")
        return

    print("Re-building the web gallery (gallery.html)...")
    try:
        subprocess.run(["python3", "build_gallery.py"], check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error running gallery builder: {e}")
        return

    print("\nAll merge and rebuilding operations completed successfully!")

if __name__ == "__main__":
    main()
