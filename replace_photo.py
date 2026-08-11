# replace_photo.py
import os
import shutil
import subprocess
import openpyxl

# Paths
SRC_IMAGE = "/Users/henrys/Library/Mobile Documents/com~apple~CloudDocs/Downloads/Joseph Novosad Rattlesnake Canyon Arches.jpg"
DEST_NAME = "Joseph_Novosad_Rattlesnake_Canyon_Arches_h.jpg"

TEMP_DIR = "/Users/henrys/source/colorado_the_beautiful/temp_images"
LOCAL_COPY_DIR = "/Users/henrys/source/colorado_the_beautiful/local copy of images"

EXCEL_PATH = "/Users/henrys/source/colorado_the_beautiful/Outreach list.xlsx"
ROW_INDEX = 35 # Row where Henry Stiles' Maroon Bells was located

def main():
    # 1. Verify source image exists
    if not os.path.exists(SRC_IMAGE):
        print(f"Error: Source image not found at: {SRC_IMAGE}")
        return

    os.makedirs(TEMP_DIR, exist_ok=True)
    os.makedirs(LOCAL_COPY_DIR, exist_ok=True)

    temp_dest = os.path.join(TEMP_DIR, DEST_NAME)
    local_copy_dest = os.path.join(LOCAL_COPY_DIR, DEST_NAME)

    # 2. Scale & Optimize photo using macOS sips
    # -Z 1200: constraints largest dimension to 1200px
    # -s formatOptions 80: sets JPEG quality to 80%
    print(f"Scaling and compressing image to temp_images...")
    sips_cmd = [
        "sips",
        "-Z", "1200",
        "-s", "formatOptions", "80",
        SRC_IMAGE,
        "--out", temp_dest
    ]
    try:
        subprocess.run(sips_cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        print("  -> Success!")
    except subprocess.CalledProcessError as e:
        print(f"  -> Error scaling image: {e}")
        return

    # 3. Update the Excel Spreadsheet Row 35
    print(f"Updating spreadsheet at row {ROW_INDEX}...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['SUBMISSION Yeses']

    # Update columns:
    # A (1): Name of Submitter
    # B (2): Type
    # C (3): Place Name
    # D (4): Place City
    # E (5): URL of Photo
    # F (6): Story
    # G (7): Category
    ws.cell(row=ROW_INDEX, column=1, value="Joseph Novosad")
    ws.cell(row=ROW_INDEX, column=2, value="Friend/Family")
    ws.cell(row=ROW_INDEX, column=3, value="Rattlesnake Canyon Arches")
    ws.cell(row=ROW_INDEX, column=4, value="Fruita, CO")
    ws.cell(row=ROW_INDEX, column=5, value=f"https://environmentamerica.org/wp-content/uploads/2026/08/{DEST_NAME}")
    ws.cell(row=ROW_INDEX, column=6, value="Rattlesnake Canyon Arches—hiking to explore McInnis Canyons National Conservation Area")
    ws.cell(row=ROW_INDEX, column=7, value="Friends and family")

    wb.save(EXCEL_PATH)
    print("  -> Spreadsheet updated and saved!")

    # 4. Run EXIF metadata tagger script
    print("Running write_exif_metadata.py script to tag the photo...")
    try:
        subprocess.run(["python3", "write_exif_metadata.py"], check=True)
        print("  -> Photo metadata tagged successfully!")
    except subprocess.CalledProcessError as e:
        print(f"  -> Error running EXIF tagger: {e}")
        return

    # 5. Copy the fully tagged photo to the local backup directory
    print(f"Copying tagged image to local copy directory...")
    shutil.copy2(temp_dest, local_copy_dest)
    print("  -> Done!")

    # 6. Run HTML compiler script to update index.html
    print("Running build_gallery.py script to rebuild index.html...")
    try:
        subprocess.run(["python3", "build_gallery.py"], check=True)
        print("  -> Gallery index.html updated successfully!")
    except subprocess.CalledProcessError as e:
        print(f"  -> Error running gallery builder: {e}")
        return

    # 7. Update Category Statistics sheet formulas
    print("Running add_stats_sheet.py script to refresh Category Statistics...")
    try:
        subprocess.run(["python3", "add_stats_sheet.py"], check=True)
        print("  -> Category statistics sheet refreshed successfully!")
    except subprocess.CalledProcessError as e:
        print(f"  -> Error updating statistics: {e}")
        return

    print("\nAll steps of the photo replacement procedure completed successfully!")

if __name__ == "__main__":
    main()
