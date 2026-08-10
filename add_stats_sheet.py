# add_stats_sheet.py
import os
import openpyxl

EXCEL_PATH = "/Users/henrys/source/colorado_the_beautiful/Outreach list.xlsx"
STATS_SHEET_NAME = "Category Statistics"

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel file not found at {EXCEL_PATH}")
        return

    print(f"Loading workbook: {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # Overwrite if sheet already exists
    if STATS_SHEET_NAME in wb.sheetnames:
        print(f"Removing existing '{STATS_SHEET_NAME}' sheet...")
        wb.remove(wb[STATS_SHEET_NAME])
        
    print(f"Creating new '{STATS_SHEET_NAME}' sheet...")
    ws = wb.create_sheet(title=STATS_SHEET_NAME)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15

    # 1. Write Headers
    ws['A1'] = "Contributor Category"
    ws['B1'] = "Tally"
    ws['C1'] = "Percentage"
    
    # Apply bold styling to headers
    bold_font = openpyxl.styles.Font(bold=True)
    ws['A1'].font = bold_font
    ws['B1'].font = bold_font
    ws['C1'].font = bold_font

    # 2. Write Category names
    categories = [
        "Friends and family",
        "Legislators",
        "Coalition partners",
        "Professional photographers",
        "Unclassified (Blank)"
    ]
    
    for idx, cat in enumerate(categories, start=2):
        ws.cell(row=idx, column=1, value=cat)
        
    # Write Total row label
    ws.cell(row=7, column=1, value="Total Submissions").font = bold_font

    # 3. Write dynamic Excel formulas for tallies (Column B)
    # B2 to B5: COUNTIF in Column G of main sheet
    ws['B2'] = "=COUNTIF('SUBMISSION Yeses'!G$2:G$1000, A2)"
    ws['B3'] = "=COUNTIF('SUBMISSION Yeses'!G$2:G$1000, A3)"
    ws['B4'] = "=COUNTIF('SUBMISSION Yeses'!G$2:G$1000, A4)"
    ws['B5'] = "=COUNTIF('SUBMISSION Yeses'!G$2:G$1000, A5)"
    
    # B6: Total active names in Column A minus the sum of categorized items
    ws['B6'] = "=COUNTA('SUBMISSION Yeses'!A$2:A$1000) - SUM(B$2:B$5)"
    
    # B7: Total sum
    ws['B7'] = "=SUM(B$2:B$6)"
    ws['B7'].font = bold_font

    # 4. Write dynamic Excel formulas for percentages (Column C)
    # C2 to C6: category tally / total tally
    for r in range(2, 7):
        ws[f'C{r}'] = f"=B{r}/B$7"
        ws[f'C{r}'].number_format = '0.0%'
        
    # C7: Total sum of percentages (should equal 100%)
    ws['C7'] = "=SUM(C$2:C$6)"
    ws['C7'].number_format = '0.0%'
    ws['C7'].font = bold_font

    # Save workbook
    wb.save(EXCEL_PATH)
    print(f"Successfully added dynamic stats sheet to: {EXCEL_PATH}")

if __name__ == "__main__":
    main()
