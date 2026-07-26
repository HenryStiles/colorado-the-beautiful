#!/usr/bin/env python3
"""
Colorado the Beautiful - Broken Link Fixer
Checks all image links in the spreadsheet, detects 404s,
resolves the correct WordPress filenames using redirects,
updates the spreadsheet, and rebuilds the gallery.
"""

import os
import re
import urllib.request
import urllib.parse
import openpyxl

EXCEL_PATH = "/Users/henrys/source/colorado_the_beautiful/Outreach list.xlsx"
BASE_IMAGE_URL = "https://environmentamerica.org/wp-content/uploads/2026/07/"

def check_url_status(url):
    """Checks the URL HTTP status and returns the final redirected URL if successful."""
    # Standard user-agent to bypass simple bot blocks
    req = urllib.request.Request(
        url, 
        method='HEAD',
        headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)'}
    )
    try:
        with urllib.request.urlopen(req, timeout=5) as response:
            return response.status, response.geturl()
    except urllib.error.HTTPError as e:
        return e.code, None
    except Exception as e:
        return 999, None

def get_base_filename(filename):
    """Strips size suffixes like -1024x768 from the filename."""
    name, ext = os.path.splitext(filename)
    # Match pattern -1024x768 at the end of the name
    name_clean = re.sub(r'-\d+x\d+$', '', name)
    return name_clean + ext

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel file not found at {EXCEL_PATH}")
        return

    print("=" * 80)
    print(" COLORADO THE BEAUTIFUL - BROKEN LINK FIXER")
    print(" Checking links and resolving WordPress redirects...")
    print("=" * 80)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['SUBMISSION Yeses']
    
    updated_count = 0
    checked_count = 0
    broken_count = 0
    
    for row_idx in range(2, 133):
        filename_val = ws.cell(row=row_idx, column=5).value
        place_name = ws.cell(row=row_idx, column=3).value
        
        if not filename_val:
            continue
            
        filename = str(filename_val).strip()
        checked_count += 1
        
        # Build direct URL
        encoded_filename = urllib.parse.quote(filename)
        url = f"{BASE_IMAGE_URL}{encoded_filename}"
        
        status, final_url = check_url_status(url)
        
        if status == 200:
            # URL works!
            continue
            
        print(f"\nRow {row_idx}: Broken Link Detected! ('{filename}' for '{place_name}')")
        print(f"  -> HTTP Status: {status}")
        broken_count += 1
        
        # Try to resolve redirect by stripping the size suffix
        base_filename = get_base_filename(filename)
        if base_filename != filename:
            print(f"  -> Stripped size suffix: '{base_filename}'")
            encoded_base = urllib.parse.quote(base_filename)
            base_url = f"{BASE_IMAGE_URL}{encoded_base}"
            
            # Follow redirects to see if WordPress sends us to the correct file
            # We use GET instead of HEAD here because some servers behave differently on redirects for GET
            get_req = urllib.request.Request(
                base_url,
                headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)'}
            )
            try:
                with urllib.request.urlopen(get_req, timeout=5) as response:
                    resolved_url = response.geturl()
                    resolved_filename = os.path.basename(urllib.parse.unquote(resolved_url))
                    
                    if response.status == 200 and resolved_filename != base_filename:
                        print(f"  🎉 RESOLVED: Redirected to working file: '{resolved_filename}'")
                        ws.cell(row=row_idx, column=5).value = resolved_filename
                        updated_count += 1
                        continue
            except Exception as e:
                print(f"  -> Failed to resolve base URL redirect: {e}")
                
        # Fallback: Try appending '-1' to the size suffix filename
        name, ext = os.path.splitext(filename)
        variant_filename = f"{name}-1{ext}"
        encoded_variant = urllib.parse.quote(variant_filename)
        variant_url = f"{BASE_IMAGE_URL}{encoded_variant}"
        
        v_status, _ = check_url_status(variant_url)
        if v_status == 200:
            print(f"  🎉 RESOLVED: Found working variant with '-1': '{variant_filename}'")
            ws.cell(row=row_idx, column=5).value = variant_filename
            updated_count += 1
            continue
            
        # Fallback 2: Try appending '-scaled' to the base filename
        base_name, base_ext = os.path.splitext(base_filename)
        scaled_filename = f"{base_name}-scaled{base_ext}"
        encoded_scaled = urllib.parse.quote(scaled_filename)
        scaled_url = f"{BASE_IMAGE_URL}{encoded_scaled}"
        
        s_status, _ = check_url_status(scaled_url)
        if s_status == 200:
            print(f"  🎉 RESOLVED: Found working scaled variant: '{scaled_filename}'")
            ws.cell(row=row_idx, column=5).value = scaled_filename
            updated_count += 1
            continue
            
        print(f"  ❌ UNRESOLVED: Could not find working server filename for '{filename}'")

    print("\n" + "=" * 80)
    print(" SCANNING & REPAIR SUMMARY")
    print("=" * 80)
    print(f"Checked:  {checked_count} image entries")
    print(f"Broken:   {broken_count} links found")
    print(f"Repaired: {updated_count} spreadsheet entries")
    
    if updated_count > 0:
        wb.save(EXCEL_PATH)
        print(f"\nSaved changes to {EXCEL_PATH}!")
        
        # Trigger gallery rebuild
        print("Rebuilding gallery.html...")
        import subprocess
        subprocess.run(["python3", "/Users/henrys/source/colorado_the_beautiful/build_gallery.py"])
    else:
        print("\nNo spreadsheet updates were needed.")

if __name__ == "__main__":
    main()
