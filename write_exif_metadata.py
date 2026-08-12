# write_exif_metadata.py
import os
import sys
import subprocess
import openpyxl

# --- Configuration ---
EXCEL_PATH = "/Users/henrys/source/colorado_the_beautiful/Outreach list.xlsx"
IMAGE_DIR = "/Users/henrys/source/colorado_the_beautiful/temp_images"
EXIFTOOL_PATH = "/opt/homebrew/bin/exiftool"
PROJECT_PREFIX = "Colorado is Beautiful 150th Anniversary"

def tag_image_metadata(filepath, author, title, story):
    """Runs exiftool to write EXIF, IPTC, and XMP metadata tags to an image file."""
    # Prepend project name before story
    if story:
        full_description = f"{PROJECT_PREFIX} - {story}"
    else:
        full_description = PROJECT_PREFIX

    # Extract organization from parentheticals or dashes if present
    import re
    org = "Environment Colorado"
    photographer = author
    m = re.search(r'\((.*?)\)', author)
    if m:
        org = m.group(1).strip()
        photographer = re.sub(r'\(.*?\)', '', author).strip()
    elif " - " in author:
        parts = author.split(" - ")
        org = parts[0].strip()
        photographer = parts[1].strip()
    elif " – " in author:
        parts = author.split(" – ")
        org = parts[0].strip()
        photographer = parts[1].strip()

    cmd = [
        EXIFTOOL_PATH,
        # 1. Author / Photographer tags
        f"-Artist={photographer}",
        f"-By-line={photographer}",
        f"-Creator={photographer}",
        
        # 2. Credit / Source tags (Organization) - Clear both to test removing the slash concatenation
        "-Credit=",
        "-Source=",
        
        # 3. License tags
        "-Copyright=Used by permission",
        "-CopyrightNotice=Used by permission",
        "-Rights=Used by permission",
        
        # 4. Story / Description tags (with project prefix)
        f"-ImageDescription={full_description}",
        f"-Caption-Abstract={full_description}",
        f"-Description={full_description}",
        
        # 5. Title / Object Name tags
        f"-ObjectName={title}",
        f"-Title={title}",
        
        # Overwrite the file in-place without creating a '_original' backup file
        "-overwrite_original",
        
        filepath
    ]
    
    try:
        subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        return True
    except subprocess.CalledProcessError as e:
        print(f"      Error tagging {os.path.basename(filepath)}: {e.stderr.decode().strip()}")
        return False

def read_exif_tags(filepath):
    """Reads and displays written EXIF tags for verification."""
    cmd = [
        EXIFTOOL_PATH,
        "-Artist", "-Creator", "-Title", "-ObjectName", "-ImageDescription", "-Caption-Abstract", "-Copyright",
        filepath
    ]
    try:
        res = subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        return res.stdout
    except Exception as e:
        return f"Error reading tags: {e}"

def main():
    test_mode = "--test" in sys.argv or "-t" in sys.argv

    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel file not found at {EXCEL_PATH}")
        return
        
    if not os.path.exists(IMAGE_DIR):
        print(f"Error: Local image directory '{IMAGE_DIR}' not found.")
        return

    print(f"Reading spreadsheet: {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['SUBMISSION Yeses']
    
    tagged_count = 0
    missing_count = 0

    rows_to_process = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        submitter = str(row[0]).strip() if row[0] else 'Anonymous'
        place_name = str(row[1]).strip() if row[1] else 'Colorado Landmark'
        filename = str(row[3]).strip() if row[3] else ''
        story = str(row[4]).strip() if row[4] else ''
        
        if not filename:
            continue
            
        rows_to_process.append((idx, submitter, place_name, filename, story))

    if test_mode:
        print(f"\n🧪 RUNNING IN TEST MODE (Processing 1 test image)...\n")
        rows_to_process = rows_to_process[:1]

    for idx, submitter, place_name, filename, story in rows_to_process:
        local_path = os.path.join(IMAGE_DIR, filename)
        
        print(f"Row {idx}: Submitter='{submitter}', Place='{place_name}', Filename='{filename}'")
        
        if os.path.exists(local_path):
            print(f"  -> Tagging EXIF metadata...")
            success = tag_image_metadata(
                filepath=local_path,
                author=submitter,
                title=place_name,
                story=story
            )
            if success:
                tagged_count += 1
                if test_mode:
                    print("\n🔍 EXIF Tags Written to Test File:")
                    print("-" * 50)
                    print(read_exif_tags(local_path))
                    print("-" * 50)
        else:
            print(f"  -> [Skip] Local file not found in '{IMAGE_DIR}'.")
            missing_count += 1
            
    print("\nMetadata tagging summary:")
    print(f"  Successfully tagged: {tagged_count} files")
    if missing_count > 0:
        print(f"  Local files missing: {missing_count} files")

if __name__ == "__main__":
    main()
