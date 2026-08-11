# add_category_column.py
import os
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

EXCEL_PATH = "/Users/henrys/source/colorado_the_beautiful/Outreach list.xlsx"

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel file not found at {EXCEL_PATH}")
        return

    print(f"Loading workbook: {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['SUBMISSION Yeses']

    # Column F (6) is "Category" in the new master sheet structure
    # Ensure header is written
    ws.cell(row=1, column=6, value="Category")
    
    # Define keywords for legislators (case-insensitive)
    legislator_keywords = ["representative", "senator", "congressman", "congresswoman", "rep.", "sen."]

    auto_legislator_count = 0
    already_filled_count = 0
    blank_count = 0

    print("Analyzing rows and applying legislator auto-classifications to empty category cells...")
    
    # Iterate through rows starting from row 2
    for r_idx in range(2, ws.max_row + 1):
        name_val = ws.cell(row=r_idx, column=1).value
        category_val = ws.cell(row=r_idx, column=6).value
        
        # Skip if name is empty
        if not name_val:
            continue
            
        name_str = str(name_val).lower().strip()
        
        # If category is already filled, keep it as is
        if category_val:
            already_filled_count += 1
            continue
            
        # Check for Legislator keywords
        if any(kw in name_str for kw in legislator_keywords):
            ws.cell(row=r_idx, column=6, value="Legislators")
            auto_legislator_count += 1
        else:
            blank_count += 1

    # Create the Data Validation dropdown menu for Column F (Column 6)
    print("Creating Excel dropdown validation for Column F...")
    options = '"Friends and family,Legislators,Professional photographers,Coalition partners"'
    
    dv = DataValidation(
        type="list",
        formula1=options,
        allow_blank=True,
        error="Your entry is not in the list",
        errorTitle="Invalid Entry",
        prompt="Please select a category from the dropdown menu",
        promptTitle="Category Selection"
    )
    
    # Add validation to the sheet
    ws.add_data_validation(dv)
    
    # Apply to F2 through F1000 (Column F)
    dv.add("F2:F1000")

    # Save workbook
    wb.save(EXCEL_PATH)
    
    print("\nSummary of categories:")
    print(f"  - Auto-classified as Legislators: {auto_legislator_count}")
    print(f"  - Retained already filled categories: {already_filled_count}")
    print(f"  - Left blank for manual selection: {blank_count}")
    print(f"\nSuccessfully saved changes to: {EXCEL_PATH}")

if __name__ == "__main__":
    main()
