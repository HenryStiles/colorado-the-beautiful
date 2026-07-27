# verify_image_specs.py
import os
import openpyxl
from PIL import Image

EXCEL_PATH = "/Users/henrys/source/colorado_the_beautiful/Outreach list.xlsx"
IMAGE_DIR = "/Users/henrys/source/colorado_the_beautiful/temp_images"

# Specifications
MIN_WIDTH = 1600
MIN_HEIGHT = 900
MAX_FILE_SIZE_MB = 15.0
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024

def check_image(filepath):
    """Inspects a single image file for size and dimension compliance."""
    if not os.path.exists(filepath):
        return {
            'status': 'MISSING',
            'error': 'File not found'
        }
    
    file_size_bytes = os.path.getsize(filepath)
    file_size_mb = file_size_bytes / (1024 * 1024)
    
    try:
        with Image.open(filepath) as img:
            width, height = img.size
            img_format = img.format
    except Exception as e:
        return {
            'status': 'CORRUPT',
            'error': str(e),
            'size_mb': file_size_mb
        }
    
    # Check constraints
    size_pass = file_size_bytes <= MAX_FILE_SIZE_BYTES
    
    # Standard check: width >= 1600 and height >= 900
    strict_dim_pass = (width >= MIN_WIDTH) and (height >= MIN_HEIGHT)
    
    # Orientation-aware check (long edge >= 1600, short edge >= 900)
    orientation_dim_pass = (max(width, height) >= MIN_WIDTH) and (min(width, height) >= MIN_HEIGHT)
    
    issues = []
    if not size_pass:
        issues.append(f"File size exceeds 15 MB limit ({file_size_mb:.2f} MB)")
    if not strict_dim_pass:
        if orientation_dim_pass:
            issues.append(f"Portrait/orientation notice: {width}x{height}px (long edge {max(width, height)}px >= 1600, short edge {min(width, height)}px >= 900)")
        else:
            issues.append(f"Dimensions below spec: {width}x{height}px (requires at least 1600px wide and 900px tall)")
            
    status = 'PASS' if (size_pass and orientation_dim_pass) else ('WARN' if (size_pass and not strict_dim_pass and orientation_dim_pass) else 'FAIL')
    
    return {
        'status': status,
        'width': width,
        'height': height,
        'size_mb': file_size_mb,
        'format': img_format,
        'strict_dim_pass': strict_dim_pass,
        'orientation_dim_pass': orientation_dim_pass,
        'size_pass': size_pass,
        'issues': issues
    }

def main():
    print("=" * 80)
    print(" COLORADO THE BEAUTIFUL - IMAGE COMPLIANCE VERIFIER")
    print(" Specs: Min Dimensions: 1600px wide x 900px tall | Max Size: 15.0 MB")
    print("=" * 80)

    # 1. Read spreadsheet entries
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Excel file not found at {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['SUBMISSION Yeses']

    spreadsheet_entries = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        submitter = str(row[0]).strip() if row[0] else 'Anonymous'
        place_name = str(row[2]).strip() if row[2] else 'Colorado Landmark'
        city = str(row[3]).strip() if row[3] else ''
        filename = str(row[4]).strip() if row[4] else ''
        story = str(row[5]).strip() if row[5] else ''

        if not filename:
            continue

        spreadsheet_entries.append({
            'row': idx,
            'submitter': submitter,
            'place': place_name,
            'city': city,
            'filename': filename
        })

    print(f"\nLoaded {len(spreadsheet_entries)} image submissions from spreadsheet.\n")

    passed_count = 0
    warn_count = 0
    fail_count = 0
    missing_count = 0
    size_exceeded_list = []
    under_sized_list = []

    print("-" * 80)
    print(f"{'Row':<5} | {'Filename':<35} | {'Dimensions':<12} | {'Size (MB)':<10} | {'Status'}")
    print("-" * 80)

    for item in spreadsheet_entries:
        row_num = item['row']
        filename = item['filename']
        filepath = os.path.join(IMAGE_DIR, filename)

        result = check_image(filepath)

        if result['status'] == 'MISSING':
            missing_count += 1
            fail_count += 1
            print(f"{row_num:<5} | {filename[:35]:<35} | {'MISSING':<12} | {'N/A':<10} | ❌ MISSING")
            continue
        elif result['status'] == 'CORRUPT':
            fail_count += 1
            print(f"{row_num:<5} | {filename[:35]:<35} | {'CORRUPT':<12} | {result['size_mb']:<10.2f} | ❌ CORRUPT ({result['error']})")
            continue

        dim_str = f"{result['width']}x{result['height']}"
        size_str = f"{result['size_mb']:.2f} MB"

        if result['status'] == 'PASS':
            passed_count += 1
            print(f"{row_num:<5} | {filename[:35]:<35} | {dim_str:<12} | {size_str:<10} | ✅ PASS")
        elif result['status'] == 'WARN':
            warn_count += 1
            print(f"{row_num:<5} | {filename[:35]:<35} | {dim_str:<12} | {size_str:<10} | ⚠️ PORTRAIT ({result['issues'][0]})")
        else:
            fail_count += 1
            print(f"{row_num:<5} | {filename[:35]:<35} | {dim_str:<12} | {size_str:<10} | ❌ FAIL: {', '.join(result['issues'])}")

        if not result['size_pass']:
            size_exceeded_list.append((filename, result['size_mb']))

        if not result['orientation_dim_pass']:
            under_sized_list.append((filename, result['width'], result['height']))

    print("\n" + "=" * 80)
    print(" SUMMARY REPORT")
    print("=" * 80)
    print(f"Total Submissions Evaluated: {len(spreadsheet_entries)}")
    print(f"  ✅ Compliant (Meeting/Exceeding 1600x900 & <= 15MB): {passed_count}")
    print(f"  ⚠️ Portrait Orientation (High-res, long edge >= 1600px): {warn_count}")
    print(f"  ❌ Failed Specs / Issues Found: {fail_count}")

    if missing_count > 0:
        print(f"\n⚠️ Missing Files ({missing_count}):")
        for item in spreadsheet_entries:
            fp = os.path.join(IMAGE_DIR, item['filename'])
            if not os.path.exists(fp):
                print(f"  - Row {item['row']}: {item['filename']} ('{item['place']}')")

    if size_exceeded_list:
        print(f"\n❌ Files Exceeding 15.0 MB Limit ({len(size_exceeded_list)}):")
        for fn, size in size_exceeded_list:
            print(f"  - {fn}: {size:.2f} MB")

    if under_sized_list:
        print(f"\n❌ Files Below 1600x900px Dimensions ({len(under_sized_list)}):")
        for fn, w, h in under_sized_list:
            print(f"  - {fn}: {w}x{h}px")

if __name__ == "__main__":
    main()
